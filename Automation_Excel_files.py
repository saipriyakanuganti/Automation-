import openpyxl as xl
from openpyxl.chart import PieChart, Reference # here from openpyxl module which contains the chart module we are importing the barchart and reference classes

def process_workbook(filename):
    wb = xl.load_workbook(filename) # used to load a excel file
    sheet = wb['housing'] # to access the specified sheet
    # cell = sheet['a1'] # to access the specified cell
    # print(cell.value) # prints the cell value
    # print(sheet.max_row) # prints the total rows in the sheet

    for row in range(2, sheet.max_row + 1): # sheet.max_row + 1 includes the total no of rows in the sheet + 1 since the range doesn't count the last number
        cell = sheet.cell(row, 9)
        original_price = cell.value
        corrected_price = cell.value * 0.2 # multiplies the value with 0.9
        corrected_price_cell = sheet.cell(row, 11) # creates a new column
        corrected_price_cell.value = original_price + corrected_price  # values are inserted in to the new column

    values = Reference(sheet,
                min_row=2,
                max_row=sheet.max_row,
                min_col=11,
                max_col=11) # for chart we are using the row from 2 to the last row in the sheet and the 4th column
    chart = PieChart() # here we are choosing the barchart
    chart.add_data(values) # here we are adding values for the barchart
    sheet.add_chart(chart, 'm2') # here we are adding the chart to the sheet in the f2 coordinate in the sheet

    wb.save('new_' + filename) # the corrected values are stored in a new file


process_workbook('housing.xlsx')
